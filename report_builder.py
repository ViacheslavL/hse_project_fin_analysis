import openpyxl
from typing import List, Tuple

from openpyxl.styles import PatternFill
from openpyxl.chart import LineChart, Series, Reference, BarChart


class DDMReportBuilder:
    def __init__(self, results: List[Tuple], output_file: str):
        wb = openpyxl.Workbook()
        idx = 0
        wb.remove(wb.active)
        for (result, scenario) in results:
            self.create_result_sheet(wb, result, scenario, idx)
            idx += 1
        wb.save(output_file)

    def create_result_sheet(self, wb: openpyxl.Workbook, scenario_result, scenario, scenario_number):
        headers = ["Ticker", "Name", "Beta", "R_e", "Market Price", "DDM Price", "Error"]
        sheet = wb.create_sheet(f"Scenario {scenario_number}")
        sheet.append(headers)
        lower_border = 0.5
        high_border = 2
        good_results = 0

        idx = 0
        for i, v in scenario_result.items():
            idx += 1
            ddm_price = v.get("ddm_price")
            close_price = v["close"]
            sheet.append([i, v["name"], v.get("beta"), v.get("r_e"), close_price, ddm_price, v.get("error")])
            if ddm_price:
                ratio = ddm_price / close_price
                if 1 <= ratio < high_border or lower_border <= ratio <= 1:
                    good_results += 1
                    for cell in sheet.iter_cols(min_col=0, max_col=7, min_row=idx + 1, max_row=idx + 1):
                        cell[0].fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            if i in scenario.get("advanced", []):
                self.build_advanced_report_for_ticker(wb, scenario_result, i, scenario)

        sheet.append([])
        sheet.cell(1, len(headers) + 2).value = "Good results:"
        sheet.cell(1, len(headers) + 3).value = good_results
        row = 2
        for k, v in scenario.items():
            if k == "advanced":
                continue
            sheet.cell(row, len(headers) + 2).value = k
            sheet.cell(row, len(headers) + 3).value = v
            row += 1

        dims = {}
        for row in sheet.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            sheet.column_dimensions[col].width = value

    def build_advanced_report_for_ticker(self, wb: openpyxl.Workbook, scenario_result, ticker: str, scenario):
        periods = scenario["total_periods"]
        result = scenario_result.get(ticker)
        sheet = wb.create_sheet(ticker)

        total_revenue_hist = result["total_revenue_hist"][::-1]
        total_revenue_growth = result["total_revenue_growth"][::-1]
        total_revenue_growth_mean = result["total_revenue_growth_mean"]
        hist_data_len = len(total_revenue_hist)
        categories = [x for x in range(-hist_data_len + 1, periods + 1)]
        sheet.append([" "] + categories)
        sheet.append(["total_revenue"] + total_revenue_hist + result["total_revenue_forward"])
        sheet.append(["total_revenue_growth"] + total_revenue_growth + [total_revenue_growth_mean]*periods)
        sheet.append([""])
        sheet.append(["prices_t"] + result["prices_t"])
        chart = LineChart()
        chart.y_axis.title = "Total Revenue"
        chart.x_axis.title = "Year"
        chart.height = 12
        chart.width = 24
        revenue_data = Reference(sheet, min_col=2, min_row=2, max_row=2, max_col=hist_data_len + periods + 1)
        cat_data = Reference(sheet, min_col=2, min_row=1, max_row=1, max_col=hist_data_len + periods + 1)
        chart.add_data(revenue_data, from_rows=True)
        sheet.add_chart(chart, "F8")
        chart.set_categories(cat_data)

        bar_chart = BarChart()
        bar_chart.y_axis.title = "Total Revenue Growth Rate"
        bar_chart.x_axis.title = "Year"
        bar_chart.height = 12
        bar_chart.width = 24
        growth_data = Reference(sheet, min_col=2, min_row=3, max_row=3, max_col=hist_data_len + periods + 1)
        bar_chart.add_data(growth_data, from_rows=True)
        bar_chart.set_categories(cat_data)
        sheet.add_chart(bar_chart, "S8")
        for i in range(scenario["total_periods"]):
            pass



