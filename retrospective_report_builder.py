from typing import List

import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import PatternFill

from ddm_retrospective import FAIR_PRICE_DATA_WINDOW


class RetrospectiveReportBuilder:
    def __init__(self, results, output_file):
        self.output_file = output_file
        wb = openpyxl.Workbook()
        idx = 0
        main_sheet = wb.active
        main_sheet.title = "Main"
        for result in results:
            name = result.data["name"]
            ticker = result.data["ticker"]
            main_sheet.cell(1, 4).value = "DDM_sum_squared"
            main_sheet.cell(1, 5).value = "Close_sum_squared"
            main_sheet.cell(idx + 2, 1).value = idx + 1
            if result.error:
                main_sheet.cell(idx + 2, 2).value = name
                main_sheet.cell(idx + 2, 3).value = result.error
                continue

            main_sheet.cell(idx + 2, 2).value = f'=HYPERLINK("{output_file}#{ticker}!A1","{name}")'
            self.create_result_sheet(wb, result)
            fair = []
            ddm = []
            hist = []
            for r in result.results:
                fair_price = r.get("fair_price")
                if fair_price is not None:
                    fair.append(fair_price)
                    ddm.append(r["ddm_price"])
                    hist.append(r["close"])
            ddm_squared = 0
            close_squared = 0
            for i in range(len(fair)):
                ddm_squared += (ddm[i] - fair[i])**2
                close_squared += (hist[i] - fair[i])**2
            main_sheet.cell(idx + 2, 4).value = ddm_squared
            main_sheet.cell(idx + 2, 5).value = close_squared
            if ddm_squared < close_squared:
                for cell in main_sheet.iter_cols(min_col=0, max_col=7, min_row=idx + 2, max_row=idx + 2):
                    cell[0].fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            idx += 1
        wb.save(output_file)

    def get_year_quotes(self, quotes):
        year_quotes = []
        for i in range(len(quotes)):
            if i % 12 == 0:
                year_quotes.append(quotes[i])
        return year_quotes

    def draw_line_chart(self, sheet, length, rows: List[int], y_title, x_title, categories, place):
        chart = LineChart()
        chart.y_axis.title = y_title
        chart.x_axis.title = x_title
        chart.height = 12
        chart.width = 24
        for r in rows:
            ref = Reference(sheet, min_col=1, min_row=r, max_row=r, max_col=length + 1)
            chart.add_data(ref, from_rows=True, titles_from_data=True)
        sheet.add_chart(chart, place)
        chart.set_categories(categories)

    def create_result_sheet(self, wb, result):
        company = result.data
        company_name = company["ticker"]
        sheet = wb.create_sheet(company_name)
        company_close_historical = self.get_year_quotes(company["close"])[::-1]
        company_close_ddm = []
        company_close_fair_price = []
        for r in result.results:
            company_close_ddm.append(r["ddm_price"])
            company_close_fair_price.append(r.get("fair_price"))
        earliest_result = result.results[-1]

        total_revenue_hist = company["total_revenue"][::-1]
        company_close_fair_price = company_close_fair_price[::-1]

        total_revenue_data_len = len(total_revenue_hist)
        close_data_len = len(company_close_historical)

        company_close_ddm += self.get_year_quotes(earliest_result["close_hist"])
        company_close_ddm = company_close_ddm[::-1]

        if total_revenue_data_len > close_data_len:
            company_close_historical = [None] * (total_revenue_data_len-close_data_len) + company_close_historical
            company_close_ddm = [None] * (total_revenue_data_len-close_data_len) + company_close_ddm
        else:
            company_close_historical = company_close_historical[abs(close_data_len-total_revenue_data_len):]
            company_close_ddm = company_close_ddm[abs(close_data_len - total_revenue_data_len):]

        diff_size = abs(total_revenue_data_len - len(company_close_fair_price))
        company_close_fair_price = [None] * diff_size + company_close_fair_price

        year = 2021
        categories = [x for x in range(-total_revenue_data_len + year, year)]
        sheet.append([" "] + categories)
        sheet.append(["Close Historical"] + company_close_historical)
        sheet.append(["Close Model"] + company_close_ddm)
        sheet.append(["Total Revenue Historical"] + total_revenue_hist)
        sheet.append(["Total Revenue Model"] + [' ' for x in earliest_result["total_revenue_hist"][::-1]] + earliest_result["total_revenue_forward"])
        sheet.append(["Close Fair Price"] + company_close_fair_price)

        cell = sheet.cell(10, 1)
        cell.value = f'=HYPERLINK("{self.output_file}#Main!A1","Back to main")'

        chart = LineChart()
        chart.y_axis.title = "Price"
        chart.x_axis.title = "Year"
        chart.height = 12
        chart.width = 24
        close_hist_data = Reference(sheet, min_col=1, min_row=2, max_row=2, max_col=close_data_len + 1)
        close_model_data = Reference(sheet, min_col=1, min_row=3, max_row=3, max_col=close_data_len + 1)
        close_fair_price = Reference(sheet, min_col=1, min_row=6, max_row=6, max_col=close_data_len + 1)
        cat_data = Reference(sheet, min_col=2, min_row=1, max_row=1, max_col=close_data_len + 1)
        chart.add_data(close_hist_data, from_rows=True, titles_from_data=True)
        chart.add_data(close_model_data, from_rows=True, titles_from_data=True)
        chart.add_data(close_fair_price, from_rows=True, titles_from_data=True)
        sheet.add_chart(chart, "F8")
        chart.set_categories(cat_data)

        cat_data = Reference(sheet, min_col=2, min_row=1, max_row=1, max_col=total_revenue_data_len + 1)
        self.draw_line_chart(sheet, total_revenue_data_len, [4,5], "Total Revenue", "Year", cat_data, "F30")


